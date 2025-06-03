import os

from aiogram import Router, F
from aiogram.types import Message
import uuid

from openpyxl import load_workbook

from app.handlers.base_handler import start_command
from app.update_file import sort_message
from database.banks import create_bank, get_amount_by_user, get_capacity_by_bank
from database.files import insert_file_name, get_bank_id_by_file_id
from database.questions import insert_questions_bulk
from database.usage_types import get_user_type
from database.users import create_user_if_not_exists, get_user_by_id


from dotenv import load_dotenv
load_dotenv()

uploading_file_router = Router()

def count_questions_in_excel(file_path: str) -> int:
    """
    Excel fayl ichidagi toâ€˜liq va toâ€˜gâ€˜ri formatlangan savollar sonini hisoblaydi.

    Tafsilotlar:
        2-qatordan boshlab 5 ustunlik qatorlarni tekshiradi.

        Barcha kataklar matn boâ€˜lishi va boâ€˜sh boâ€˜lmasligi kerak.

        Xatolik yuz bersa, 0 qaytaradi
    :param file_path:
    :return: savollar soni
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        return sum(
            1 for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True)
            if all(row) and all(isinstance(cell, str) and cell.strip() for cell in row)
        )


    except Exception as e:
        print(f"Xatolik yuz berdi: {e}")
        return 0


def extract_questions_from_excel(file_path: str) -> list:
    """
    Excel fayldan barcha savollarni (savol matni + 1 toâ€˜gâ€˜ri + 3 notoâ€˜gâ€˜ri javob) lugâ€˜at (dict) sifatida ajratib oladi.

    Tafsilotlar:
        Har bir savol dictionary koâ€˜rinishida: {"question": ..., "correct": ..., "wrong1": ..., "wrong2": ..., "wrong3": ...}

        Xatolik yuz bersa, boâ€˜sh roâ€˜yxat qaytaradi.
    :param file_path:
    :return: savol va variantlarni qaytaradi
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        return [
            {
                "question": row[0].strip(),
                "correct": row[1].strip(),
                "wrong1": row[2].strip(),
                "wrong2": row[3].strip(),
                "wrong3": row[4].strip()
            }
            for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True)
            if all(row) and all(isinstance(cell, str) and cell.strip() for cell in row)
        ]


    except Exception as e:
        print(f"Xatolik yuz berdi: {e}")
        return []


async def create_new_bank(user_id: int, title: str, file_path: str) -> int | None:
    """
    Yangi test banki yaratadi, unga savollarni yuklaydi va bank_id ni qaytaradi.

    Tafsilotlar:
        Bank foydalanuvchiga bogâ€˜lanadi (user_id)

        Savollarni file_path dagi fayldan oladi

        Agar savollar mavjud boâ€˜lsa bank_id, aks holda None qaytaradi

    :param user_id:
    :param title:
    :param file_path:
    :return: Integer yoki None
    """
    bank_id = (await create_bank(user_id, title))["bank_id"]
    questions = extract_questions_from_excel(file_path)
    await insert_questions_bulk(bank_id, questions)
    return bank_id if len(questions) != 0 else None


async def save_uploaded_file(file, bot) -> str:
    """
    Foydalanuvchi yuborgan faylni saqlaydi.

    Tafsilotlar:
        Fayl nomiga maxsus id qoâ€˜shiladi (takrorlanishni oldini olish uchun)

        Asl fayl turi saqlanadi (.xlsx)

        Faylni bot yordamida yuklab olib, bazaga yozadi

        Saqlangan faylning toâ€˜liq manzilini (path) qaytaradi

    :param file:
    :param bot:
    :return: fayl yo'lini qaytaradi
    """
    suffix = file.file_name.rfind(".")
    file_name = f"{file.file_name[:-5]}_{uuid.uuid4()}{file.file_name[suffix - 1:]}"
    file_name.replace(" ", "_")
    file_path = f"downloads/{file_name}"
    file_content = await bot.download(file.file_id)

    with open(file_path, "wb") as f:
        f.write(file_content.read())

    return file_path


async def check_count_questions_in_excel(msg: Message, file_path):
    count = count_questions_in_excel(file_path)
    if count == 0:
        await start_command(msg, "âŒ Faylda hech qanday savol topilmadi. Iltimos, faylni tekshirib qayta yuboring.")
        os.remove(file_path)
        return True
    return False


async def process_valid_excel_file(message, file_path: str, original_file_name: str):
    """
    Yaroqli faylni tahlil qilib, undagi savollarni maâ€™lumotlar bazasiga yozadi.

    Tafsilotlar:
        Savollar sonini hisoblab, foydalanuvchiga javob yuboradi

        Foydalanuvchini roâ€˜yxatdan oâ€˜tkazadi (agar mavjud boâ€˜lmasa)

        Bank yaratadi va unga savollarni bogâ€˜laydi

        Fayl nomini bank bilan bogâ€˜lab saqlaydi

    :param message:
    :param file_path:
    :param original_file_name:
    :return:
    """

    if await check_count_questions_in_excel(message, file_path):
        return
    user_id = message.from_user.id
    chat_id = message.chat.id
    username = message.from_user.username or ""
    await create_user_if_not_exists(user_id=user_id, chat_id=chat_id, username=username)

    bank_name = original_file_name.replace(".xlsx", "")
    bank_id = await create_new_bank(user_id, bank_name, file_path)
    await insert_file_name(bank_id, file_path)

    await start_command(message, f"âœ… Savollar bazaga qoâ€˜shildi.")


def check_excel_file(file_path: str):
    try:
        workbook = load_workbook(file_path)
        sheetnames = workbook.sheetnames

        if len(sheetnames) > 1:
            text = (
                "âš ï¸ <b>Diqqat!</b>\n\n"
                "Faylingizda faqat bitta <i>aktiv sahifa</i> boâ€˜lishi kerak! ğŸ¯\n\n"
                "ğŸ—‚ï¸ Agar bir nechta sahifa (sheet) boâ€˜lsa, bot faqat birinchi sahifani oâ€˜qiydi.\n"
                "Shuningdek, boshqa sahifalar eâ€™tibordan chetda qoladi.\n\n"
                "âœ… Iltimos, faqat bitta sahifani qoldirib, qolganlarini olib tashlang va qayta yuboring."

            )
            return False, text

        sheet = workbook.active
        real_question_count = sum(
            1 for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True)
            if row[0] and str(row[0]).strip()
        )

        if real_question_count > 500:
            return False, "âš ï¸ Diqqat! Faylingizda 500 tadan koâ€˜p savol bor. Iltimos, faqat keraklilarni yuboring."

        return True, f"âœ… Fayl yuklandi. {real_question_count} ta savol bor."

    except Exception as e:
        return False, f"ğŸ˜• Xatolik yuz berdi: {str(e)}"


# foydalanuvchilarga qo'yilgan limitlarni tekshirish
# admin uchun tekshirish o'tqizilmaydi
async def check_user_limit(message: Message):
    file = message.document

    if not file.file_name.endswith(".xlsx"):
        return False, "âŒ Iltimos, faqat .xlsx formatdagi fayl yuboring."

    user_id = message.from_user.id
    check = await get_user_by_id(user_id)

    if check is None:
        return (True,)

    check_caption = True if message.caption is None else False
    if not check_caption:
        file_id = message.caption
        bank_id = (await get_bank_id_by_file_id(file_id))[0]["bank_id"]
        if len(bank_id) == 0:
            return False, "âŒ Bunday ID li fayl topilmadi."
        # else:
        #     bank_id = bank_id[0][0]
        capacity = (await get_capacity_by_bank(bank_id))[0]["capacity"]
        amount = 0
    else:
        amount = (await get_amount_by_user(user_id))[0]["count"]
        capacity = 1
    usage_type = (await get_user_type(user_id))[0]["usage_type"]
    if (amount >= 3 or capacity <= 0) and usage_type == "ordinary":
        if amount >= 3:
            return False, "âŒ Testlar bazangizda joy qolmagan."
        else:
            return False, "âŒ Bu testni qayta yangilash imkoniyatingiz tugadi."
    elif (amount >= 5 or capacity <= 0) and usage_type == "pro":
        if amount >= 5:
            return False, "âŒ Testlar bazangizda joy qolmagan."
        else:
            return False, "âŒ Bu testni qayta yangilash imkoniyatingiz tugadi."
    elif usage_type not in {"founder", "ordinary", "pro"}:
        return False, "âš ï¸ Tizimda xatolik yuz berdi. Iltimos, adminlar bilan bog'laning."



    return (True,)

@uploading_file_router.message(F.document)
async def handle_excel_file(message: Message):
    """
    Foydalanuvchi yuborgan .xlsx faylni tekshiradi va agar hammasi toâ€˜gâ€˜ri boâ€˜lsa, bazaga yozishni boshlaydi.

    Tafsilotlar:
        Fayl formatini tekshiradi (.xlsx boâ€˜lishi shart)

        Faylni yuklab oladi va validatsiyadan oâ€˜tkazadi (check_excel_file)

        Xatolik boâ€˜lsa, foydalanuvchini ogohlantiradi va faylni oâ€˜chiradi

        Validatsiyadan oâ€˜tsa, process_valid_excel_file chaqiriladi

    :param message:
    :return:
    """
    file = message.document
    # faylni update qilish kerak
    check_limit = await check_user_limit(message)
    if not check_limit[0]:
        await start_command(message, check_limit[1])
        return

    file_path = await save_uploaded_file(file, message.bot)

    is_valid, result_message = check_excel_file(file_path)
    if not is_valid:
        await start_command(message, f"{result_message}")
        os.remove(file_path)
        return

    await message.answer(f"âœ… {result_message}\n\nğŸ”„ Bazaga yozilmoqda...")

    if not message.caption is None:
        await sort_message(message, file_path)
        return

    await process_valid_excel_file(message, file_path, file.file_name)
